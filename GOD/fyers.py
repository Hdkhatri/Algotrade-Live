import credentials as c
from fyers_apiv3 import fyersModel
import pandas as pd
import webbrowser
import time
import pyperclip
from datetime import datetime, timedelta
from fyers_apiv3.FyersWebsocket import data_ws
import os
from openpyxl import Workbook, load_workbook
import logging

# Configure logging
logging.basicConfig(
    filename=os.path.join('Log', 'live.log'),
    level=logging.INFO,
    format='%(asctime)s %(levelname)s: %(message)s'
)

# Replace these values with your actual API credentials
client_id = c.client_id
secret_key = c.secret_key
redirect_uri = c.redirect_uri
response_type = c.response_type 
state = c.state
user_name = c.user_name
grant_type = c.grant_type  


# Create a session model with the provided credentials
session = fyersModel.SessionModel(
    client_id=client_id,
    secret_key=secret_key,
    redirect_uri=redirect_uri,
    response_type=response_type
)

def create_fyers_session(authcode):
    """
    Create a session object to handle the Fyers API authentication and token generation.
    Returns an authenticated FyersModel instance.
    """
    session = fyersModel.SessionModel(
        client_id=client_id,
        secret_key=secret_key, 
        redirect_uri=redirect_uri, 
        response_type=response_type, 
        grant_type=grant_type
    )
    session.set_token(authcode)
    response = session.generate_token()
    access_token = response['access_token']
    fyers_main = fyersModel.FyersModel(client_id=client_id, is_async=False, token=access_token, log_path="")
    return fyers_main,access_token

def fetch_and_save_history(access_token, SYMBOL, TIMEFRAME, FROM_DATE, TO_DATE):
    """
    Fetch historical data from Fyers API and save as CSV.
    Returns the CSV filename.
    """

    data = {
        "symbol": SYMBOL,
        "resolution": TIMEFRAME,
        "date_format": "1",
        "range_from": FROM_DATE,
        "range_to": TO_DATE,
        "cont_flag": "1"
    }
    fyers = fyersModel.FyersModel(client_id=client_id, is_async=False, token=access_token, log_path="")
    response = fyers.history(data=data)
    print("Response:", response)
    df = pd.DataFrame(response['candles'], columns=['time','open','high','low','close','volume'])
    print(df.iloc[-1]['close'])
    df['datetime'] = pd.to_datetime(df['time'],unit='s',utc=True).map(lambda x: x.tz_convert('Asia/Kolkata'))
    csv_name = "RawData/"+data['symbol'].split(":")[1]+str(data['resolution'])+"history"+data['range_from']+"_"+data['range_to']+".csv"
    print(csv_name)
    df.to_csv(csv_name, header=True, index=True)
    return csv_name

def fetch_data(access_token, SYMBOL, TIMEFRAME, FROM_DATE, TO_DATE):
    """
    Fetch historical data from Fyers API and save as CSV.
    Returns the CSV filename.
    """

    data = {
        "symbol": SYMBOL,
        "resolution": TIMEFRAME,
        "date_format": "1",
        "range_from": FROM_DATE,
        "range_to": TO_DATE,
        "cont_flag": "1"
    }
    fyers = fyersModel.FyersModel(client_id=client_id, is_async=False, token=access_token, log_path="")
    response = fyers.history(data=data)
    if 'candles' in response:
        df = pd.DataFrame(response['candles'], columns=['time','open','high','low','close','volume'])
        #print(df.iloc[-1]['close'])
        df['datetime'] = pd.to_datetime(df['time'],unit='s',utc=True).map(lambda x: x.tz_convert('Asia/Kolkata'))
        return df
    else:
        print("Error fetching data:", response)
        return pd.DataFrame()
        


def add_signals_and_strikes(df):
    # Clean column names
    df.columns = df.columns.str.strip()

    # Ensure correct data type
    df['close'] = pd.to_numeric(df['close'], errors='coerce')

    # Drop rows where 'close' is NaN
    df.dropna(subset=['close'], inplace=True)

    # === Compute EMA 11 and EMA 17 ===
    df['EMA_11'] = df['close'].ewm(span=11, adjust=False).mean()
    df['EMA_17'] = df['close'].ewm(span=17, adjust=False).mean()

    # === Initialize Signal and Strike columns ===
    df['Signal'] = None
    df['Strike_Price'] = None

    # === Position state tracker ===
    position_state = 0  # 0 = neutral, 1 = long, -1 = short
    Strike_gap = 1000  # Gap for strike price calculation
    # === Logic to generate buy/sell signals ===
    for i in range(len(df)):
        close = df.loc[i, 'close']
        ema11 = df.loc[i, 'EMA_11']
        ema17 = df.loc[i, 'EMA_17']

        # Flip to long
        if close > ema11 and close > ema17 and position_state != 1:
            df.at[i, 'Signal'] = 'Sell Short'
            nearest_500 = round(close / 1000) * 1000
            put_strike = nearest_500 - Strike_gap
            df.at[i, 'Strike_Price'] = f'SELL {put_strike} PE'
            position_state = 1

        # Flip to short
        elif close < ema11 and close < ema17 and position_state != -1:
            df.at[i, 'Signal'] = 'Sell Long'
            nearest_500 = round(close / 1000) * 1000
            call_strike = nearest_500 + Strike_gap
            df.at[i, 'Strike_Price'] = f'SELL {call_strike} CE'
            position_state = -1

    return df

def get_NextToNextweekThursday(date_str):
    dt = datetime.strptime(date_str, "%Y-%m-%d")
    # Find this week's Thursday
    days_to_this_thursday = (3 - dt.weekday() + 7) % 7
    this_week_thursday = dt + timedelta(days=days_to_this_thursday)
    # Move to next week's Thursday
    next_week_thursday = this_week_thursday + timedelta(days=7)
    # Move to next to next week's Thursday
    next_to_nextweek_thursday = next_week_thursday + timedelta(days=7)
    return next_to_nextweek_thursday.strftime("%Y-%m-%d")



def getSymbol(Ex_UnderlyingSymbol, expiry_date, strike, opt_type, Ex="NSE"):
    # Check if expiry_date is the last Thursday of the month
    dt_expiry = datetime.strptime(expiry_date, "%Y-%m-%d")
    # Find last day of the month
    next_month = dt_expiry.replace(day=28) + timedelta(days=4)
    last_day_of_month = next_month - timedelta(days=next_month.day)
    # Find last Thursday of the month
    days_to_last_thursday = (last_day_of_month.weekday() - 3) % 7
    last_thursday = last_day_of_month - timedelta(days=days_to_last_thursday)
    YY = str(dt_expiry.year)[2:]
    strike_str = str(int(strike))
    opt_type_str = "CE" if opt_type.lower() == "call" else "PE"
    if dt_expiry == last_thursday:
        MMM = dt_expiry.strftime("%b").upper()
        # DYNAMICALLY HANDLE MONTHLY EXPIRY FOR NIFTY OPTIONS
        # Uncomment the next line if you want to use dynamic month handling
        symbol = f"{Ex}:{Ex_UnderlyingSymbol}{YY}{MMM}{strike_str}{opt_type_str}"
        # if opt_type.lower() == "call":
        #     # MANUALLY handle JUL expiry for NIFTY options
        #     symbol = f"{Ex}:{Ex_UnderlyingSymbol}25DEC27000{opt_type_str}"  
        # else:
        #     # MANUALLY handle JUL expiry for NIFTY options
        #     symbol = f"{Ex}:{Ex_UnderlyingSymbol}25DEC21000{opt_type_str}"  
        # MANULY handle JUL expiry for NIFTY options
        #symbol = f"{Ex}:{Ex_UnderlyingSymbol}25DEC{strike_str}{opt_type_str}"
        return symbol
    # expiry_date in format YYYY-MM-DD
    dt = datetime.strptime(expiry_date, "%Y-%m-%d")
    YY = str(dt.year)[2:]
    M = str(dt.month)
    # Use special month codes for October, November, December
    if dt.month == 10:
        M = "O"
    elif dt.month == 11:
        M = "N"
    elif dt.month == 12:
        M = "D"
    dd = str(dt.day)
    strike_str = str(int(strike))
    opt_type_str = "CE" if opt_type.lower() == "call" else "PE"
    # DYNAMICALLY HANDLE WEEKLY EXPIRY FOR NIFTY OPTIONS
    # Uncomment the next line if you want to use dynamic week handling
    symbol = f"{Ex}:{Ex_UnderlyingSymbol}{YY}{M}{dd}{strike_str}{opt_type_str}"

    # MANULY handle JUL expiry for NIFTY options
    # if opt_type.lower() == "call":
    #         # MANUALLY handle JUL expiry for NIFTY options
    #         symbol = f"{Ex}:{Ex_UnderlyingSymbol}25DEC27000{opt_type_str}"  
    # else:
    #         # MANUALLY handle JUL expiry for NIFTY options
    #         symbol = f"{Ex}:{Ex_UnderlyingSymbol}25DEC21000{opt_type_str}"  
    # #symbol = f"{Ex}:{Ex_UnderlyingSymbol}25DEC{strike_str}{opt_type_str}"
    return symbol





def ema_ribbon_strategy(df, len1=8, len2=20, len3=4, len4=8, len5=20):
    """
    Applies EMA Ribbon strategy and returns DataFrame with signal and strike label.

    Parameters:
        df (pd.DataFrame): Must contain ['time', 'open', 'high', 'low', 'close', 'volume', 'datetime']
        len1, len2, ..., len5 (int): EMA lengths

    Returns:
        pd.DataFrame: Original df + 'EMA1', 'EMA2', ..., 'Signal', 'Strike_Label'
    """

    # Copy DataFrame to avoid modifying original
    df = df.copy()
    Strike_gap = 1000  # Gap for strike price calculation   
    # === Calculate EMAs ===
    df['EMA1'] = df['close'].ewm(span=len1, adjust=False).mean()
    df['EMA2'] = df['close'].ewm(span=len2, adjust=False).mean()
    df['EMA3'] = df['close'].ewm(span=len3, adjust=False).mean()
    df['EMA4'] = df['close'].ewm(span=len4, adjust=False).mean()
    df['EMA5'] = df['close'].ewm(span=len5, adjust=False).mean()

    # === Identify Signal Conditions ===
    buy_condition = (df['EMA1'] > df['EMA2']) & (df['EMA1'].shift(1) <= df['EMA2'].shift(1))
    sell_condition = (df['EMA1'] < df['EMA2']) & (df['EMA1'].shift(1) >= df['EMA2'].shift(1))

    # === Initialize Signal Column ===
    df['Signal'] = ''
    df.loc[buy_condition, 'Signal'] = 'Sell Short'
    df.loc[sell_condition, 'Signal'] = 'Sell Long'

    # === Strike Price Calculation ===
    df['Nearest500'] = (df['close'] / 1000).round() * 1000
    df['PutStrike'] = df['Nearest500'] - Strike_gap
    df['CallStrike'] = df['Nearest500'] + Strike_gap

    # === Generate Strike Label based on Signal ===
    df['Strike_Price'] = ''
    df.loc[df['Signal'] == 'Sell Short', 'Strike_Price'] = 'SELL ' + df['PutStrike'].astype(int).astype(str) + ' PE'
    df.loc[df['Signal'] == 'Sell Long', 'Strike_Price'] = 'SELL ' + df['CallStrike'].astype(int).astype(str) + ' CE'

    return df

def ema_breakout_strategy(df):
    """
    Applies EMA 20 High/Low breakout strategy.
    Assumes df has columns: time, open, high, low, close, volume, datetime
    Returns a new DataFrame with signals and strike labels
    """
    df = df.copy()
    
    # Calculate EMA of high and low (20-period)
    df['ema_high'] = df['high'].ewm(span=20, adjust=False).mean()
    df['ema_low'] = df['low'].ewm(span=20, adjust=False).mean()
    
    # Initialize columns for signal and label
    df['Signal'] = None
    df['Label'] = None
    
    position = None  # 'long' or 'short' or None

    for i in range(1, len(df)):
        close0 = df.at[i, 'close']
        close1 = df.at[i-1, 'close']
        emah0 = df.at[i, 'ema_high']
        emah1 = df.at[i-1, 'ema_high']
        emal0 = df.at[i, 'ema_low']
        emal1 = df.at[i-1, 'ema_low']
        
        # Calculate nearest strike
        nearest_500 = round(close0 / 500) * 500
        long_strike = nearest_500 - 1000
        short_strike = nearest_500 + 1000

        # Long entry condition (last 2 closes above ema_high)
        if close0 > emah0 and close1 > emah1 and position is None:
            df.at[i, 'Signal'] = 'Long'
            df.at[i, 'Strike_Price'] = f'SELL {long_strike} PE'
            position = 'long'

        # Short entry condition (last 2 closes below ema_low)
        elif close0 < emal0 and close1 < emal1 and position is None:
            df.at[i, 'Signal'] = 'Short'
            df.at[i, 'Strike_Price'] = f'SELL {short_strike} CE'
            position = 'short'

        # Exit condition: Opposite signal
        elif position == 'long' and close0 < emal0 and close1 < emal1:
            df.at[i, 'Signal'] = 'Exit Long'
            position = None

        elif position == 'short' and close0 > emah0 and close1 > emah1:
            df.at[i, 'Signal'] = 'Exit Short'
            position = None

    return df


class FyersLiveData:
    def __init__(self, access_token, symbols, data_type="SymbolUpdate", log_path="", litemode=False, write_to_file=False, reconnect=True):
        self.access_token = access_token
        self.symbols = symbols
        self.data_type = data_type
        self.log_path = log_path
        self.litemode = litemode
        self.write_to_file = write_to_file
        self.reconnect = reconnect
        self.fyers = None

    def on_message(self, message):
        print("Response:", message)

    def on_error(self, message):
        print("Error:", message)

    def on_close(self, message):
        print("Connection closed:", message)

    def on_open(self):
        self.fyers.subscribe(symbols=self.symbols, data_type=self.data_type)
        self.fyers.keep_running()

    def connect(self):
        self.fyers = data_ws.FyersDataSocket(
            access_token=self.access_token,
            log_path=self.log_path,
            litemode=self.litemode,
            write_to_file=self.write_to_file,
            reconnect=self.reconnect,
            on_connect=self.on_open,
            on_close=self.on_close,
            on_error=self.on_error,
            on_message=self.on_message
        )
        self.fyers.connect()


def create_xlsx_file(symbol):
            """
            Creates an empty XLSX file in the 'Data' folder with the given symbol as filename.
            Returns the file path.
            """
            SYMBOL = symbol.split(":")[1]  # Extract the symbol part after 'NSE:'
            print(f"Creating xlsx file for symbol: {SYMBOL}")
            try:
                SYMBOL = symbol.split(":")[1]  # Extract the symbol part after 'NSE:'
                folder = "Data"
                file_path = os.path.join(folder, f"{SYMBOL}.xlsx")
                if not os.path.exists(folder):
                    os.makedirs(folder)
                    wb = Workbook()
                    ws = wb.active
                    ws.append(['OPEN', 'HIGH', 'LOW', 'CLOSE', 'DATETIME', 'EMA1', 'EMA2', 'EMA3', 'EMA4', 'EMA5', 'SIGNAL', 'STRIKE_PRICE'])
                    wb.save(file_path)

                return file_path
            except Exception as e:
                print(f"Error creating xlsx file: {e}")
                return None
class FyersLiveData:
    def __init__(self, fyers, symbol):
        self.fyers = fyers
        self.symbol = symbol
        
    def get_ltp(self):
        try:
            data = self.fyers.quotes({"symbols": self.symbol})
            ltp = data['d'][0]['v']['lp']
            return ltp
        except Exception as e:
            print(f"Error getting LTP for {self.symbol}: {str(e)}")
            raise


def calculate_live_signals(file_path, current_close):
        """
        Calculate EMAs and signals for live data using previous data points.
        Updates the Excel file with new signals.
        
        Parameters:
            file_path (str): Path to the Excel file
            current_close (float): Current closing price
        """
        try:
            logging.info(f"Calculating signals for close price: {current_close}")
            
            # Read existing data from Excel
            df = pd.read_excel(file_path)
            
            # Add new row for current data
            new_row = pd.DataFrame({
                'CLOSE': [current_close],
                'DATETIME': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            })
            df = pd.concat([df, new_row], ignore_index=True)
            
            # Calculate EMAs
            df['EMA1'] = df['CLOSE'].ewm(span=8, adjust=False).mean()
            df['EMA2'] = df['CLOSE'].ewm(span=13, adjust=False).mean()
            df['EMA3'] = df['CLOSE'].ewm(span=21, adjust=False).mean()
            df['EMA4'] = df['CLOSE'].ewm(span=34, adjust=False).mean()
            df['EMA5'] = df['CLOSE'].ewm(span=55, adjust=False).mean()

            # Get last row for signal calculation
            last_row = df.iloc[-1]
            prev_row = df.iloc[-2] if len(df) > 1 else None
            
            # Initialize signal variables
            signal = None
            strike_price = None
            option_symbol = None
            Strike_gap = 500
            
            # Calculate signal based on EMAs
            if prev_row is not None:
                close = last_row['CLOSE']
                ema1 = last_row['EMA1']
                ema2 = last_row['EMA2']
                
                # Previous values
                prev_ema1 = prev_row['EMA1']
                prev_ema2 = prev_row['EMA2']
                
                # Get next-to-next Thursday for expiry
                today = datetime.now().strftime("%Y-%m-%d")
                expiry_date = get_NextToNextweekThursday(today)
                
                # Signal conditions with option symbol generation
                if ema1 > ema2 and prev_ema1 <= prev_ema2:
                    signal = 'Sell Short'
                    nearest_500 = round(close / 500) * 500
                    put_strike = nearest_500 - Strike_gap
                    strike_price = f'SELL {put_strike} PE'
                    
                    # Get option symbol for PUT
                    option_symbol = getSymbol(
                        Ex_UnderlyingSymbol="NIFTY",
                        expiry_date=expiry_date,
                        strike=put_strike,
                        opt_type="put"
                    )
                    
                elif ema1 < ema2 and prev_ema1 >= prev_ema2:
                    signal = 'Sell Long'
                    nearest_500 = round(close / 500) * 500
                    call_strike = nearest_500 + Strike_gap
                    strike_price = f'SELL {call_strike} CE'
                    
                    # Get option symbol for CALL
                    option_symbol = getSymbol(
                        Ex_UnderlyingSymbol="NIFTY",
                        expiry_date=expiry_date,
                        strike=call_strike,
                        opt_type="call"
                    )

            # Get the last row's values
            result = {
                'EMA1': last_row['EMA1'],
                'EMA2': last_row['EMA2'],
                'EMA3': last_row['EMA3'],
                'EMA4': last_row['EMA4'],
                'EMA5': last_row['EMA5'],
                'Signal': signal,
                'Strike_Price': strike_price,
                'Option_Symbol': option_symbol
            }
            
            if signal:
                logging.info(f"Signal generated: {signal} | Strike: {strike_price} | Symbol: {option_symbol}")
            
            return result
            
        except Exception as e:
            logging.error(f"Error calculating live signals: {str(e)}")
            return None



def store_ltp_data(ltp_list, file_path):
    """
    Store LTP data with OHLC calculations in Excel file.
    
    Parameters:
        ltp_list (list): List of LTP values
        file_path (str): Path to the Excel file
    """
    try:
        logging.info(f"Processing LTP data for storage in {file_path}")
        
        if not ltp_list:
            logging.error("Empty LTP list provided")
            return False

        # Calculate OHLC values
        open_price = ltp_list[0]
        high_price = max(ltp_list)
        low_price = min(ltp_list)
        close_price = ltp_list[-1]
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Load existing workbook
        if os.path.exists(file_path):
            wb = load_workbook(file_path)
            ws = wb.active
        else:
            logging.info("Creating new Excel file")
            wb = Workbook()
            ws = wb.active
            # Add headers if new file
            ws.append(['OPEN', 'HIGH', 'LOW', 'CLOSE', 'DATETIME', 'EMA1', 'EMA2', 'EMA3', 'EMA4', 'EMA5', 'SIGNAL', 'STRIKE_PRICE', 'OPTION_SYMBOL'])

        # Calculate signals and EMAs
        signals = calculate_live_signals(file_path, close_price)
        
        if signals is None:
            return False

        # Append new data
        ws.append([
            open_price,
            high_price,
            low_price,
            close_price,
            current_time,
            signals['EMA1'],
            signals['EMA2'],
            signals['EMA3'],
            signals['EMA4'],
            signals['EMA5'],
            signals['Signal'],
            signals['Strike_Price'],
            signals['Option_Symbol']
        ])

        # Save the workbook
        wb.save(file_path)
        logging.info(f"Successfully stored OHLC data in {file_path}")
        
        # If there's a signal, log it prominently
        if signals['Signal']:
            logging.info(f"*** NEW SIGNAL: {signals['Signal']} - {signals['Strike_Price']} | Symbol: {signals['Option_Symbol']} ***")
        
        return True

    except Exception as e:
        logging.error(f"Error storing LTP data: {str(e)}")
        return False


